import logging
logger = logging.getLogger(__name__)
import sys
import argparse
import json
import socket
import time
import re
import pkgutil
from pydoc import safeimport
import os
import boto3
import json
import string
import botocore.exceptions
import botocore.errorfactory
import lens
import workload
import openpyxl
from openpyxl import Workbook, load_workbook

# set logging temporarily before args parsing
loglevel = os.getenv("LOG_CFG", "ERROR")
logger = logging.getLogger(__name__)
logging.basicConfig(level=loglevel)

# handle arguments

if __name__ == '__main__':
    argdesc = sys.argv[0]
else:
    argdesc = __name__

parser = argparse.ArgumentParser(description=argdesc, formatter_class=argparse.RawDescriptionHelpFormatter)
parser.add_argument("command", nargs=1, choices=["gen", "parse"], help="The action to take; gen-quare generates an XLS using name, parse-quare parses XLS using name")
parser.add_argument("xls_file", nargs='?', default="stdout", help="The name of the XLS file associated with command.")
parser.add_argument("--loglevel", choices=["DEBUG", "WARNING", "INFO", "ERROR", "CRITICAL"], default="ERROR", help="set logger debug level; default is ERROR")
parser.add_argument("--wlname", nargs="?", help="Use named workload to gen or parse.")
#parser.add_argument("--output", choices=["default","signalk"], default="default", help="specify format of JSON written to stdout; default is 'default'")
#parser.add_argument("--format", choices=["compact","pretty"], default="compact", help="specify whether output should be pretty-printed; default is 'compact'")
args = parser.parse_args()
#logging.debug("Arguments = %s" % args)

# Remove all handlers associated with the root logger object, to set final format and log level
for handler in logging.root.handlers[:]:
    logging.root.removeHandler(handler)

logging.basicConfig(format='%(name)s %(levelname)s:%(asctime)s %(message)s',datefmt='%m/%d/%Y %I:%M:%S',level=args.loglevel)
logger = logging.getLogger(__name__)
logger.debug("Arguments = %s" % args)

# generates a workload with designated name (or default) and emits PSV or an XLS file; if workload with name exists, emits the current value of the workload
if args.command[0] == "gen":
# this code assumes the CLI context is the account you want to work with
    wlname = args.wlname
    if args.wlname == None:
        wlname = workload.DEFAULT_WLNAME
    wlid = workload.get_workload(wlname=wlname)
    client = boto3.client('wellarchitected')
    wl = client.get_workload(WorkloadId=wlid)
    lenses = wl['Workload']['Lenses']

    if args.xls_file != "stdout":
        wb = Workbook()
        sh = wb.active
        wb.remove(sh)
        for l in lenses:
            sh = wb.create_sheet(l)
            lens.write_lens_xls(wlid,l,sh)
        wb.save(filename = args.xls_file)
    else:
        for l in lenses:
            lens.write_lens_psv(wlid,l)

# parses an XLS file and updates a workload in the current account
if args.command[0] == "parse":
    wlname = args.wlname
    if args.wlname == None:
        wlname = workload.DEFAULT_WLNAME
    wlid = workload.get_workload(wlname=wlname,gen=False)
    if wlid == "":
        logger.error("Unable to find specified workload: ", wlname)
        sys.exit(1)
    try:
        client = boto3.client('wellarchitected')
        wl = client.get_workload(WorkloadId=wlid)
        lenses = wl['Workload']['Lenses']

        wb = load_workbook(args.xls_file)
        shlist = wb.sheetnames
        for i in shlist:
            if not i in lenses:
                continue

            sh = wb[i]
            
            # this code assumes the rows are sorted by qid
            cur_qid = None
            cur_cnone = False
            ans_list = []
            qstat = True
            reason = "NONE"
            for row in sh.iter_rows(min_row=2, min_col=1, max_row=sh.max_row+1, max_col=sh.max_column+1):
                p = row[lens.P_OFFSET].value
                qid = row[lens.QID_OFFSET].value
                cid = row[lens.CID_OFFSET].value
                ct = row[lens.CTITLE_OFFSET].value
                r = row[lens.RESP_OFFSET].value
                notes = row[lens.NOTES_OFFSET].value
                if p is None: p = ""
                if qid is None: qid = ""
                if cid is None: cid = ""
                if ct is None: ct = ""
                if r is None: r = ""
                if notes is None: notes = ""

                # update a choice
                if cur_qid is None:
                    cur_qid = qid
                if cur_qid != qid:
                    # update answers for workload
                    try:
                        if qstat == True:
                            #print("Update choices: ", wlid, i, cur_qid, ans_list, notes)
                            res = client.update_answer(WorkloadId=wlid, LensAlias=i, QuestionId=cur_qid, SelectedChoices=ans_list)
                        else:
                            res = client.update_answer(WorkloadId=wlid, LensAlias=i, QuestionId=cur_qid, SelectedChoices=[], Notes=notes, IsApplicable=False, Reason=reason)
                    except Exception as e:
                        logger.error("Error updating answer: ", qid, ans_list)
                        print(e)
                    cur_qid = qid
                    ans_list = []
                    cur_cnone = False
                    cur_qid = qid
                    qstat = True
                    reason = "NONE"
                else:
                    if cid != "":
                        if r in ["X", "x"]:
                            if cid.endswith("_no"):
                                cur_cnone = True
                                ans_list = [cid]
                            else:
                                ans_list.append(cid)
                    else:
                        if r in ["NA", "N/A", "na", "n/a", "N/a", "Na"]:
                            qstat = False
                            reason = "BUSINESS PRIORITIES"

                if logger.getEffectiveLevel() == logging.DEBUG: print("Pillar = " + p + ", Qid = " + qid + ", cid = " + cid + ", ctitle = " + ct + ", resp = " + r)

    except Exception as e:
        print(e)
        logger.error("problem parsing file")
        sys.exit(1)

sys.exit(0)
