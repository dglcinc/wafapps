import boto3
import json
import string
import botocore.exceptions
import botocore.errorfactory
import logging
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import DEFAULT_FONT

P_OFFSET = 0
QID_OFFSET = 1
CID_OFFSET = 3
CTITLE_OFFSET = 4
RESP_OFFSET = 8
NOTES_OFFSET = 9

try:
        from openpyxl.cell import get_column_letter
except ImportError:
        from openpyxl.utils import get_column_letter
        from openpyxl.utils import column_index_from_string

pillar_list = ["security", "performance", "reliability", "costOptimization", "operationalExcellence", "sustainability"]

client = boto3.client("wellarchitected")
logger = logging.getLogger('root')

def remove_non_ascii(a_str):
    ascii_chars = set(string.printable)

    return ''.join(filter(lambda x: x in ascii_chars, a_str))

def print_lens(lens_dict, f=None):
    print(json.dumps(lens_dict,indent=2,default=str), file=f)

def write_lens_psv(wl_id,lens_id,f=None,**kwargs):
    # This function writes lens's question data as pipe separated values
    if f != None:
        print("sep=|", file=f)
        print("lens|pillar id|question id|question title|choice id|choice title|choice status|choice reason|question description|X or NA|Reason/Notes", file=f)
    
    for pi in pillar_list:
        wl = client.get_workload(WorkloadId=wl_id)

        try:
            ans = client.list_answers(WorkloadId=wl_id,LensAlias=lens_id,PillarId=pi)
        except:
            logger.error("no answers found for pillar: " + pi)
            continue

        for asum in ans['AnswerSummaries']:
            qid = asum['QuestionId']
            qt = asum['QuestionTitle']
            remove_non_ascii(qt)
    
            for ch in asum['Choices']:
                ct = ch['Title']
                ct = ct.replace('\n','')
                ct = " ".join(ct.split())
                cid = ch['ChoiceId']
                cd = ch['Description']
                cd = cd.replace('\n','')
                cd = " ".join(cd.split())

                cas = list(filter(lambda summaries: summaries['ChoiceId'] == cid, asum['ChoiceAnswerSummaries']))
                reason = ""
                status = ""
                if len(cas) > 0:
                    status = cas[0]['Status']
                    reason = cas[0]['Reason']
    
                print(lens_id + "|" + pi + "|" + qid + "|" + qt + "|" + cid + "|" + ct + "|" + status + "|" + reason + "|" + cd, file=f)

def write_lens_xls(wl_id, lens_id, sh):
    # this function writes an XLS containing the lens questions so they can be reviewed and/or updated easily
    row = ["pillar", "question_id", "question_title", "choice_id", "choice_title", "choice_status", "choice_reason", "choice_description", "X or NA", "Reason/Notes"]
    sh.append(row)

    for pi in pillar_list:
        wl = client.get_workload(WorkloadId=wl_id)
        if (logger.level == "DEBUG"):
            print(json.dumps(wl,indent=2,default=str))
    
        try:
            ans = client.list_answers(WorkloadId=wl_id,LensAlias=lens_id,PillarId=pi)
        except:
            logger.error("no answers found for pillar: " + pi)
            continue
    
        if (logger.getEffectiveLevel() == logging.DEBUG):
            print(json.dumps(ans,indent=2,default=str))
        for asum in ans['AnswerSummaries']:
            qid = asum['QuestionId']
            qt = asum['QuestionTitle']
            remove_non_ascii(qt)
            ia = ""
            re = ""
            try:
                ia = asum['IsApplicable']
                re = asum['Reason']
            except:
                pass
            row = [pi, qid, qt, "", "", ia, re, "", "", ""]
            sh.append(row)
    
            for ch in asum['Choices']:
                ct = ch['Title']
                ct = ct.replace('\n','')
                ct = " ".join(ct.split())
                cid = ch['ChoiceId']
                cd = ch['Description']
                cd = cd.replace('\n','')
                cd = " ".join(cd.split())
    
                cas = list(filter(lambda summaries: summaries['ChoiceId'] == cid, asum['ChoiceAnswerSummaries']))
                reason = ""
                status = ""
                if len(cas) > 0:
                    status = cas[0]['Status']
                    reason = cas[0]['Reason']
    
                row = [pi, qid, qt, cid, ct, status, reason, cd, "", ""]

                sh.append(row)
    # put in formatting features

    # set to font calibri 14, left/top, auto-wrap
    font = Font(name='Calibri', size=14, bold=False, vertAlign=None,underline='none',strike=False,color='FF000000')
    alignment = Alignment(horizontal='left', vertical='top', text_rotation=0, wrap_text=True)

    for rows in sh.iter_rows(min_row=1, max_row=sh.max_row, min_col=1, max_col=sh.max_column):
        for cell in rows:
            cell.font = font
            cell.alignment = alignment

    # set width
    sh.column_dimensions['A'].width = 23
    sh.column_dimensions['C'].width = 30
    sh.column_dimensions['E'].width = 36
    sh.column_dimensions['H'].width = 65
    sh.column_dimensions['I'].width = 15
    sh.column_dimensions['J'].width = 40
    #for column_cells in sh.columns:
    #    new_column_length = max(len(str(cell.value)) for cell in column_cells)
    #    new_column_letter = (get_column_letter(column_cells[0].column))
    #    if new_column_length > 0:
    #        sh.column_dimensions[new_column_letter].width = new_column_length

    # hide the ID columns
    sh.column_dimensions['B'].hidden = True
    sh.column_dimensions['D'].hidden = True
    sh.column_dimensions['F'].hidden = True
    sh.column_dimensions['G'].hidden = True

    # conditional formatting for pillar and question_title (columns a and c)
    cf_list = [
        { 'range' : "A2:A" + str(sh.max_row), 'rule' : ['COUNTIF($A$2:$A2,$A2)>1'] },
        { 'range' : "C2:C" + str(sh.max_row), 'rule' : ['COUNTIF($C$2:$C2,$C2)>1'] }
        ]

    for cf in cf_list:
        sh.conditional_formatting.add(range_string=cf['range'], cfRule=FormulaRule(formula=cf['rule'],font=Font(color="FFFFFF")))
